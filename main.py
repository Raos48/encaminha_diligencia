import json
import os
import re
import time
import traceback
import warnings
from datetime import date
import logging
import sys
import subprocess
import chromedriver_autoinstaller
import urllib3
import colorama
from colorama import Fore, Back, Style
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import requests
from packaging import version
import shutil

colorama.init(autoreset=True)

# Configuração inicial
GITHUB_REPO = "Raos48/encaminha_diligencia"
VERSION = "1.0.9"


def check_for_update():
    try:
        current_version = VERSION
        print(Fore.CYAN + Style.BRIGHT + f"Versão atual: {current_version}" + Style.RESET_ALL)

        # Verificar última versão no GitHub
        api_url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        response = requests.get(api_url, timeout=10)

        if response.status_code == 200:
            latest_release = response.json()
            latest_version = latest_release['tag_name'].replace('v', '')

            if version.parse(latest_version) > version.parse(current_version):
                # Mensagem de alerta para nova versão disponível
                print(Fore.MAGENTA + Style.BRIGHT + 
                      f" ATUALIZAÇÃO DISPONÍVEL " + Style.RESET_ALL)
                print(Fore.YELLOW + Style.BRIGHT + f"Nova versão: {latest_version}" + Style.RESET_ALL)
                print(Fore.YELLOW + Style.BRIGHT + "Atualize o sistema para obter as últimas melhorias." + Style.RESET_ALL)
                input(Fore.GREEN + Style.BRIGHT + "Aperte ENTER para continuar..." + Style.RESET_ALL)

                # Escrever a informação da nova versão em um arquivo
                with open("update_available.json", 'w') as update_file:
                    json.dump({
                        "update_available": True,
                        "version": latest_version,
                        "download_url": latest_release['assets'][0]['browser_download_url']
                    }, update_file)
                return True
            else:
                print(Fore.GREEN + Style.BRIGHT + "Sistema já está na versão mais recente." + Style.RESET_ALL)
                return False
        else:
            print(Fore.RED + Style.BRIGHT + "Erro ao acessar API do GitHub para verificar atualizações." + Style.RESET_ALL)
            return False
    except Exception as e:
        print(Fore.RED + Style.BRIGHT + f"Erro ao verificar atualização: {str(e)}" + Style.RESET_ALL)
        return False

check_for_update()


warnings.filterwarnings("ignore", category=DeprecationWarning)
urllib3.disable_warnings()
chromedriver_autoinstaller.install()

logging.getLogger('urllib3').setLevel(logging.CRITICAL)  # Suprime warnings do urllib3
logging.basicConfig(stream=sys.stdout, level=print)  # Define o nível de logging global para INFO


options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')
options.add_experimental_option('excludeSwitches', ['enable-logging'])  # Suprime mensagens do Chrome

service = Service(executable_path=chromedriver_autoinstaller.install())
driver = webdriver.Chrome(options=options, service=service)
driver.maximize_window()

timeout = 30

caminho_arquivo = 'protocolos.xlsx'
workbook = openpyxl.load_workbook(caminho_arquivo)
worksheet = workbook['Plan1']

linha = 2

# Função para verificar se há uma nova versão disponível

# Verificar se há atualizações antes de iniciar o restante do código


try:
    print(Fore.YELLOW + "Iniciando Automação.." + Style.RESET_ALL)
    print(Fore.YELLOW + "Após a conclusão do processo de Login pressione ENTER" + Style.RESET_ALL)

    driver.get("https://esisrec.inss.gov.br/esisrec/pages/painel_de_entrada/consultar_painel_de_entrada.xhtml")

    while True:
        try:
            # Tenta encontrar o elemento (sem tempo limite)
            driver.find_element(By.XPATH, "/html/body/div[3]/div/fieldset/div/form/div[2]/div/div/div[1]/h4")
            print(Fore.CYAN + "Continuando..." + Style.RESET_ALL)
            break  # Sai do loop se o elemento for encontrado
        except NoSuchElementException:
            print(Fore.YELLOW + "Aguardando login..." + Style.RESET_ALL)
            time.sleep(5)  # Aguarda 5 segundos antes de tentar novamente

    while True:
        try:
            protocolo = worksheet.cell(row=linha, column=2).value
            status_coluna3 = worksheet.cell(row=linha, column=3).value
            id = worksheet.cell(row=linha, column=1).value
            
            if not protocolo:
                print(Fore.GREEN + "Finalizado!" + Style.RESET_ALL)
                input()
                break

            if status_coluna3:
                print(Fore.YELLOW + f"ID {id} já processado, pulando..." + Style.RESET_ALL)
                linha += 1
                continue
            
            print(Fore.CYAN + "======================================================================" + Style.RESET_ALL)
            print(Fore.CYAN + f"Iniciando processamento do Protocolo:id Nº{id} - {protocolo}" + Style.RESET_ALL)
            driver.get("https://esisrec.inss.gov.br/esisrec/pages/painel_de_entrada/consultar_painel_de_entrada.xhtml")
            ActionChains(driver).send_keys(Keys.HOME).perform()
            
            print(Fore.CYAN + "Busca Processo..." + Style.RESET_ALL)
            campo_busca = driver.find_element(By.ID, "formBuscaRapida:buscaRapida")
            campo_busca.clear()
            campo_busca.send_keys(protocolo)
            WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            driver.find_element(By.ID, "formBuscaRapida:btBuscaRapida").click()
            time.sleep(2)

            e_nb = driver.find_element(By.ID, "formProcesso:numeroIdentificacao").get_attribute("value")
            unidade = driver.find_element(By.ID, "formProcesso:orgaoAtual").text
            driver.execute_script("window.scrollTo(0, 1000)")
            
                    
            if len(driver.find_elements(By.ID, "formProcesso:encaminharProcesso")) == 0:
                situacao = "Não permite encaminhamento, verificar se o processo encontra-se arquivado ou no CRPS."
                print(Fore.RED + situacao + Style.RESET_ALL)
                worksheet.cell(row=linha, column=3).value = f"ERRO:{situacao}"
                workbook.save(caminho_arquivo)
                linha += 1
                continue
                
            print(Fore.CYAN + "Clicar em Analisar..." + Style.RESET_ALL)
            element = driver.find_element(By.ID, "formProcesso:analisarAnexarDocumentos")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
            time.sleep(1)
            element.click()
            time.sleep(2)
            
            print(Fore.CYAN + "Lançar Evento..." + Style.RESET_ALL)
            element = driver.find_element(By.XPATH, "//span[@id='iptEventoLancado']/button")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
            time.sleep(1)
            driver.find_element(By.XPATH, "//span[@id='iptEventoLancado']/button").click()
            time.sleep(1)
            
            print(Fore.CYAN + "Selecionar Evento Diligencia não cumprida..." + Style.RESET_ALL)
            driver.find_element(By.XPATH, "//span[@id='iptEventoLancado_panel']/ul/li[28]").click()
            
            data_formatada = date.today().strftime("%d/%m/%Y")
            driver.execute_script("scroll(0,500)")
            time.sleep(1)            
            element = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span/div[2]/div/div/div[3]/div[1]/div/div[3]/div/div/ul/li[2]/a")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)            
            driver.find_element(By.LINK_TEXT, "Digitar um Documento").click()
            time.sleep(1)
            driver.find_element(By.ID, "anexar_docs:tipoDocDigitarDocumentos_label").click()                  
            time.sleep(1)            
            print(Fore.CYAN + "Selecionar Documento tipo DESPACHO..." + Style.RESET_ALL)
            driver.find_element(By.ID, "anexar_docs:tipoDocDigitarDocumentos_40").click()
            time.sleep(5)
                        
            print(Fore.CYAN + "Redigir Despacho..." + Style.RESET_ALL)
            element = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span/div[2]/div/div/div[3]/div[1]/div/div[3]/div/div/div/div[2]/div[2]/div/div[2]/div[1]/p[5]")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)            
            
            editor = driver.find_element(By.XPATH, "//div[@id='anexar_docs:textEditor_editor']/div")
            editor.clear()
            texto_despacho = (
                f'\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t{data_formatada}\n\n\n'
                f'Protocolo: {protocolo}\nRecorrente: INSTITUTO NACIONAL DO SEGURO SOCIAL – INSS\n\n\n'
                f'\tDevolução do processo de recurso conforme solicitação do órgão julgador.'
            )
            editor.send_keys(texto_despacho)
            time.sleep(1)
            
            print(Fore.CYAN + "Anexar Documento..." + Style.RESET_ALL)
            element = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span/div[2]/div/div/div[3]/div[1]/div/div[3]/div/div/div/div[2]/div[3]/button[1]/span")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
            driver.find_element(By.ID, 'anexar_docs:digitar_doc_anexar').click()
            time.sleep(1)
            
            driver.execute_script("window.scrollTo(0, 1000)")
            print(Fore.CYAN + "Movimentar em Bloco..." + Style.RESET_ALL)            
            driver.execute_script("document.getElementById('btnMovimentarEmBloco').click();")
            time.sleep(1)
            
            print(Fore.CYAN + "Aguardar retorno..." + Style.RESET_ALL)
            WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/div[2]/div/ul/li/span")))
            time.sleep(1)
            
            mensagem_retorno = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div/ul/li/span").text
            if mensagem_retorno != "Processo movimentado com sucesso.":
                print(Fore.YELLOW + "Retorno falhou, tentando movimentar novamente..." + Style.RESET_ALL)
                driver.execute_script("document.getElementById('btnMovimentarEmBloco').click();")
            time.sleep(1)
            
            print(Fore.GREEN + "Encaminhamento bem sucedido, salvando informações na planilha..." + Style.RESET_ALL)
            worksheet.cell(row=linha, column=3).value = mensagem_retorno
            worksheet.cell(row=linha, column=4).value = driver.find_element(By.XPATH, "/html/body/div[3]/div/form/div/div[1]/div[1]/div/div[1]/div[6]/div[1]/div/span").text
            
            elemento1 = mensagem_retorno
            elemento2 = driver.find_element(By.XPATH, '/html/body/div[3]/div/form/div/div[1]/div[1]/div/div[1]/div[6]/div[1]/div/span').text
            print(Fore.CYAN + f"{elemento1} - {elemento2}" + Style.RESET_ALL)
            
            linha += 1
            workbook.save(caminho_arquivo)

        except Exception as e:
            print(Fore.RED + "\nOcorreu um erro durante a execução:" + Style.RESET_ALL)
            print(Fore.RED + f"Erro: {str(e)}" + Style.RESET_ALL)
            print(Fore.RED + f"Detalhes do erro:\n{traceback.format_exc()}" + Style.RESET_ALL)
            input(Fore.YELLOW + "\nPressione ENTER para continuar..." + Style.RESET_ALL)
            
            erro_resumido = str(e).split('\n')[0][:100]
            worksheet.cell(row=linha, column=3).value = f"ERRO: {erro_resumido}"
            workbook.save(caminho_arquivo)
            linha += 1
            continue
finally:
    driver.quit()  # Garante que o WebDriver será encerrado
